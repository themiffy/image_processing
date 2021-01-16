from openpyxl import load_workbook

class Patient:
    def __init__(self,num,excel_file = './data.xlsx'):
        excel_wb = load_workbook(excel_file)
        sheet = excel_wb['Лист1']

        self.id = sheet.cell(row = num, column = 1).value
        self.age = sheet.cell(row = num, column = 2).value
        self.menstr = sheet.cell(row = num, column = 3).value
        self.pulp = sheet.cell(row = num, column = 4).value
        self.mutation = sheet.cell(row = num, column = 5).value
        self.quality = sheet.cell(row = num, column = 6).value
        self.size = sheet.cell(row = num, column = 7).value
        self.echogene = sheet.cell(row = num, column = 8).value
        self.echostruct = sheet.cell(row = num, column = 9).value
        self.capsule = sheet.cell(row = num, column = 10).value
        self.sharpness = sheet.cell(row = num, column = 11).value
        self.str = sheet.cell(row = num, column = 12).value
        self.shape = sheet.cell(row = num, column = 13).value
        self.orient = sheet.cell(row = num, column = 14).value
        self.vascularity = sheet.cell(row = num, column = 15).value
        self.blood = sheet.cell(row = num, column = 16).value
        self.diagnosis_prediction = sheet.cell(row = num, column = 17).value
        self.diagnosis = sheet.cell(row = num, column = 18).value

arPatients = []
for i in range(2,32):
    arPatients.append(Patient(i))

quantity = 0
age = 0

for patient in arPatients:
    if patient.pulp == 1:
        quantity += 1
        age += patient.age

print("Минимальный размер до 15мм \n")
print("Количество:", quantity, "Средний возраст:", age/quantity)



